import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Any


class ExcelService:

    @staticmethod
    def load_excel_file(path: str) -> Workbook:
        """Load an Excel file and return the Workbook."""
        return openpyxl.load_workbook(path)

    @staticmethod
    def list_excel_sheets(excel_file: Workbook) -> List[str]:
        """List names of all sheets in the workbook."""
        return excel_file.sheetnames

    @staticmethod
    def get_sheet(excel_file: Workbook, sheet_name: str) -> Worksheet:
        """Return a worksheet object by name."""
        return excel_file[sheet_name]

    @staticmethod
    def list_columns_fom_sheet(sheet: Worksheet) -> List[str]:
        """
        Return a list of column headers from the first row.
        Assumes the first row contains headers.
        """
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        return headers

    @staticmethod
    def list_unique_values_from_column_of_excel_sheet(
        sheet: Worksheet, column_name: str
    ) -> List[Any]:
        """
        Return unique values from a given column by its header.
        """
        # Find the column index based on header row
        col_idx = None
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value == column_name:
                col_idx = idx
                break

        if col_idx is None:
            raise ValueError(f"Column '{column_name}' not found.")

        # Collect unique values from rows (skip header)
        values = set()
        for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell_value = row[0].value
            if cell_value is not None:
                values.add(cell_value)

        return list(values)
